import sys
from collections import defaultdict
from datetime import datetime, timedelta
from time import sleep

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from survey_studio_clients.api_clients.outgoing_calls import SurveyStudioOutgoingCallsClient

from src.params.results_rtk import RESULTS
from src.tasks.base_automation import BaseAutomation
from src.utils.get_yesterday import get_yesterday_date


class OutgoingCallsDailyReportMaker(BaseAutomation):
    PARAMS_NUMBER = 2

    def __init__(self, client: SurveyStudioOutgoingCallsClient, yesterday: datetime) -> None:
        super().__init__(client)
        self._project_id = self._get_project_id()
        self._date_from = self._get_date_as_iso_string(yesterday)

    @staticmethod
    def _show_usage_example() -> None:
        print("You have to specify both token and project ID:\n")
        print("\tpoetry run python src/tasks/get_outgoing_calls.py yourtoken123 55555")

    def _get_project_id(self) -> str:
        return sys.argv[2]

    def _get_raw_data(self) -> pd.DataFrame:
        raw_data = self._ss_client.get_dataframe(self._project_id, self._date_from, self._date_from)

        return raw_data[["Результат", "Фактический канал"]]

    @staticmethod
    def _make_maps(data: pd.DataFrame) -> tuple[defaultdict, defaultdict]:
        results_to_channel_map = defaultdict(lambda: defaultdict(int))
        channels_counter_map = defaultdict(int)

        for _, row in data.iterrows():
            result = row["Результат"]
            channel = row["Фактический канал"]
            if pd.isna(channel):
                channel = "no_name"

            results_to_channel_map[result][channel] += 1
            channels_counter_map[channel] += 1

        return results_to_channel_map, channels_counter_map

    def _make_report(self, results_to_channel_map: defaultdict, channels_counter_map: defaultdict) -> pd.DataFrame:
        rows = []

        if not channels_counter_map:
            rows.append(f"За {self._date_from} не было звонков")
            return pd.DataFrame(rows)

        channels = sorted(channels_counter_map.keys())

        header = self._get_report_header(channels)
        rows.append(header)
        rows.append([""] + ["Количество", "Процент"] * (len(channels_counter_map.keys()) + 1))
        rows.append(["Результат"])

        channels_total = 0
        results_total_over_all_channels = defaultdict(int)

        for channel in channels:
            channels_total += channels_counter_map[channel]

        for result in RESULTS:
            total_result_over_all_channels = 0

            for channel in channels:
                channel_per_result = results_to_channel_map[result].get(channel, 0)
                total_result_over_all_channels += channel_per_result

            results_total_over_all_channels[result] = total_result_over_all_channels

        for result in RESULTS:
            row = [result]

            for channel in channels:
                channels_per_result = results_to_channel_map.get(result, {})
                channel_per_result = channels_per_result.get(channel, 0)
                channel_total = channels_counter_map[channel]
                per_cent = channel_per_result / channel_total

                row += [channel_per_result, per_cent]

            row.append(str(results_total_over_all_channels[result]))
            row.append(results_total_over_all_channels[result] / channels_total)

            rows.append(row)

        row = ["Total"]
        for channel in channels:
            row.append(channels_counter_map[channel])
            row.append(1)

        row += [channels_total, 1]
        rows.append(row)

        return pd.DataFrame(rows)

    @staticmethod
    def _get_report_header(channels: list) -> list[str]:
        header = ["Фактический канал"]
        for channel in channels:
            header.append(channel)
            header.append("")

        header.append("Total")
        header.append("")

        return header

    @staticmethod
    def _make_excel_workbook(df: pd.DataFrame) -> openpyxl.Workbook:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        for _, row in df.iterrows():
            worksheet.append(row.to_list())

        worksheet.column_dimensions["A"].width = 50
        worksheet.column_dimensions["B"].width = 10
        worksheet.column_dimensions["C"].width = 10
        worksheet.column_dimensions["D"].width = 10
        worksheet.column_dimensions["E"].width = 10
        worksheet.column_dimensions["F"].width = 10
        worksheet.column_dimensions["G"].width = 10
        worksheet.column_dimensions["H"].width = 10
        worksheet.column_dimensions["I"].width = 10

        for i in range(len(df.columns)):
            if i == 0:
                continue

            next_letter_code = ord("A") + i

            if next_letter_code % 2 == 1:
                for cell in worksheet[chr(next_letter_code)]:
                    cell.number_format = "0.00%"

        alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        font = Font(color="000000", bold=True)
        fill = PatternFill(fill_type="solid", fgColor="FFFF00")

        for row_index, row in enumerate(worksheet.iter_rows()):
            if row_index in (0, 1, 2):
                for cell in row[1:]:
                    cell.font = font
                    cell.fill = fill

            for cell in row:
                cell.alignment = alignment

        return workbook

    def _get_report_file_name(self) -> str:
        file_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        return f"./reports/report_outgoing_calls_{self._date_from}_{self._project_id}_{file_date}.xlsx"

    def run(self) -> None:
        raw_data = self._get_raw_data()
        results_to_channel_map, channels_counter_map = self._make_maps(raw_data)

        report_df = self._make_report(results_to_channel_map, channels_counter_map)
        workbook = self._make_excel_workbook(report_df)

        file_name = self._get_report_file_name()
        workbook.save(file_name)
        print(f"File {file_name} has been successfully saved")


if __name__ == "__main__":
    yesterday = get_yesterday_date()
    if yesterday.weekday() == 6:  # sunday
        for delta in range(2, -1, -1):
            day = yesterday - timedelta(days=delta)
            report_maker = OutgoingCallsDailyReportMaker(SurveyStudioOutgoingCallsClient, day)
            report_maker.run()
            if delta != 0:
                print(
                    "Waiting for 62 seconds because Survey Studio API allows only one request to requestoperatorworktime per 60 seconds..."
                )
                sleep(62)
    else:
        report_maker = OutgoingCallsDailyReportMaker(SurveyStudioOutgoingCallsClient, yesterday)
        report_maker.run()
