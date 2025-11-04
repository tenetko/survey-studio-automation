import sys
from datetime import datetime, timedelta
from time import sleep

import openpyxl
import pandas as pd
import sidetable
from openpyxl.styles import Alignment, Font, PatternFill
from survey_studio_clients.api_clients.outgoing_calls import SurveyStudioOutgoingCallsClient

from src.tasks.base_automation import BaseAutomation
from src.utils.get_yesterday import get_yesterday_date


class CallsGroupsDailyReportMaker(BaseAutomation):
    PARAMS_NUMBER = 2

    def __init__(self, client: SurveyStudioOutgoingCallsClient, yesterday: datetime) -> None:
        super().__init__(client)
        self._project_id = self._get_project_id()
        self._date_from = self._get_date_as_iso_string(yesterday)

    @staticmethod
    def _show_usage_example() -> None:
        print("You have to specify both token and project ID:\n")
        print("\tpoetry run python src/tasks/get_outgoing_calls.py yourtoken123 55555")

    def _get_project_id(self) -> str:
        return sys.argv[2]

    def _get_raw_data(self) -> pd.DataFrame:
        return self._ss_client.get_dataframe(self._project_id, self._date_from, self._date_from, True)

    @staticmethod
    def _make_excel_workbook(df: pd.DataFrame) -> openpyxl.Workbook:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        for _, row in df.iterrows():
            worksheet.append(row.to_list())

        # worksheet.column_dimensions["A"].width = 50
        # worksheet.column_dimensions["B"].width = 10
        # worksheet.column_dimensions["C"].width = 10
        # worksheet.column_dimensions["D"].width = 10
        # worksheet.column_dimensions["E"].width = 10
        # worksheet.column_dimensions["F"].width = 10
        # worksheet.column_dimensions["G"].width = 10
        # worksheet.column_dimensions["H"].width = 10
        # worksheet.column_dimensions["I"].width = 10

        # for i in range(len(df.columns)):
        #     if i == 0:
        #         continue
        #
        #     next_letter_code = ord("A") + i
        #
        #     if next_letter_code % 2 == 1:
        #         for cell in worksheet[chr(next_letter_code)]:
        #             cell.number_format = "0.00%"

        # alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        # font = Font(color="000000", bold=True)
        # fill = PatternFill(fill_type="solid", fgColor="A8A8A8")

        # for row_index, row in enumerate(worksheet.iter_rows()):
        #     if row_index in (0, 1, 2):
        #         for cell in row[1:]:
        #             cell.font = font
        #             cell.fill = fill
        #
        #     for cell in row:
        #         cell.alignment = alignment

        return workbook

    def _get_report_file_name(self) -> str:
        file_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        return f"./reports/report_calls_groups_{self._date_from}_{self._project_id}_{file_date}.xlsx"

    def run(self) -> None:
        raw_data = self._get_raw_data()

        pivot = pd.pivot_table(
            raw_data,
            index=["Регион оператора связи", "Оператор связи"],
            columns=["Результат"],
            aggfunc={"Результат": "count"},
            fill_value=0,
        )

        pivot = pivot.rename(columns={"Результат": ""})
        pivot.columns = [column[0] if column[0] else column[1] for column in pivot.columns]

        pivot_total = (pivot / pivot.sum()) * 100
        pivot_stb = pivot_total.stb.subtotal(sub_label="Итог").round(2)

        file_name = self._get_report_file_name()
        # workbook = self._make_excel_workbook(pivot_stb)
        # workbook.save(file_name)

        pivot_stb.to_excel(file_name)

        print(f"File {file_name} has been successfully saved")


if __name__ == "__main__":
    yesterday = get_yesterday_date()
    if yesterday.weekday() == 6:  # sunday
        for delta in range(2, -1, -1):
            day = yesterday - timedelta(days=delta)
            report_maker = CallsGroupsDailyReportMaker(SurveyStudioOutgoingCallsClient, day)
            report_maker.run()
            if delta != 0:
                print(
                    "Waiting for 62 seconds because Survey Studio API allows only one request to requestoperatorworktime per 60 seconds..."
                )
                sleep(62)
    else:
        report_maker = CallsGroupsDailyReportMaker(SurveyStudioOutgoingCallsClient, yesterday)
        report_maker.run()
